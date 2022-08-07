import openpyxl 

book = openpyxl.Workbook() # Criar Planilha

planilha = book.worksheets[0] # Pegando 1ª sheet

# Visualizar páginas existentes na planilha
print(book.sheetnames)

# Criando uma página
book.create_sheet('Frutas')

# Selecionar uma página (sheet)
frutas_page = book['Frutas']
frutas_page.append(['Banana', '5', 'R$3,90'])
frutas_page.append(['Maça', '4', 'R$6,99'])
frutas_page.append(['Melao', '1', 'R$4,90'])

# Inserir dados aleatorios na 1ª sheet
planilha['a1'] = 'FRUTAS' # Escrevendo na célula A1
planilha['A2'] = 'Banana' 
planilha.title = 'Planilha de Frutas' 

# Salvando
book.save(r"C:\Users\lavieira\Documents\Treinamentos\venv-python\env\AquivoExcelPython.xlsx")

# ---------------------------------

# Abrir uma planilha
book2 = openpyxl.load_workbook(r"C:\Users\lavieira\Documents\Treinamentos\venv-python\env\AquivoExcelPython.xlsx")

# Selecionando uma sheet
frutas_page = book2['Frutas']

# Imprimindo dados de cada linha
for rows in frutas_page.iter_rows(min_row=0, max_row=3): # Itera sobre cada linha   
    for cell in rows: # Itera sobre cada celula
        print(cell.value) # Imprimir cada valor da celula
        if cell.value == 'Banana': # Se encontrar este valor    
            cell.value = 'Bananaa' # Substitui por este valor onde encontrou 'Banana'